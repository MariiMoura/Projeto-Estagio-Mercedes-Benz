# =============================================================
# AUTOMAÇÃO DE ANÁLISE DE PRs TIPO J — SAP GUI SCRIPTING
# Verifica se cada PR possui Elemento PEP (Investimento)
# e grava o resultado na coluna "Comentários" do Excel de saída
#
# Planilha de entrada esperada: colunas "Tipo_IO", "Nº PR", "Centro de Custo"
# Tipo J identificado por: Tipo_IO terminado em "J" (ex: "540J")
# =============================================================

import win32com.client
import pandas as pd
import pythoncom
import time
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk
import threading


# =========================================================
# ESPERA ELEMENTO SAP
# =========================================================
def wait_for_element(session, element_id, timeout=10):
    """
    Tenta localizar um elemento SAP pelo ID dentro de um timeout.

    CORREÇÃO v1: timeout era 2s — insuficiente para servidores mais lentos.
    Aumentado para 10s como padrão, evitando falsos negativos.
    """
    start = time.time()
    while time.time() - start < timeout:
        try:
            return session.findById(element_id)
        except:
            time.sleep(0.3)
    raise Exception(f"Timeout ({timeout}s): elemento não encontrado → {element_id}")


# =========================================================
# CONEXÃO SAP
# =========================================================
def get_sap_session():
    """
    Obtém a sessão ativa do SAP GUI Scripting.
    Lança exceção legível caso o SAP não esteja aberto/logado/com scripting ativo.
    """
    try:
        sap = win32com.client.GetObject("SAPGUI")
        app = sap.GetScriptingEngine
        connection = app.Children(0)
        session = connection.Children(0)
        return session
    except Exception as e:
        raise Exception(
            "\nErro ao conectar no SAP. Verifique:\n"
            "  - SAP está aberto e logado\n"
            "  - Scripting habilitado (Options > Accessibility > Enable Scripting)\n"
            f"\nErro técnico: {e}"
        )


# =========================================================
# FECHAR JANELAS MODAIS RESIDUAIS
# =========================================================
def fechar_modais(session):
    """
    Fecha qualquer wnd[1], wnd[2] etc. que possa ter ficado aberta de
    uma iteração anterior (ex: popup de erro, popup de confirmação).
    Modais bloqueiam comandos na wnd[0], então isso precisa rodar
    sempre antes de qualquer navegação na janela principal.
    """
    # Tenta fechar até 3 janelas modais empilhadas, da mais alta para a wnd[1]
    for indice in (3, 2, 1):
        try:
            session.findById(f"wnd[{indice}]")
            session.findById(f"wnd[{indice}]").sendVKey(12)  # F12 = Cancelar/Fechar
            time.sleep(0.3)
        except:
            pass  # essa janela não existe — ok, continua


# =========================================================
# VOLTAR DA ORDEM/ATRIBUIÇÕES PARA A TELA INICIAL DO ME53N
# =========================================================
def voltar_para_inicio_me53n(session):
    """
    Replica o "volta com F3" descrito no processo manual, mas de forma
    garantida: pressiona F3 em sequência (Atribuições → Ordem → PR → ME53N
    inicial) e, ao final, confirma redigitando ME53N no campo de transação.

    CAUSA RAIZ DO BUG REPORTADO:
    O código anterior nunca dava F3 dentro de verificar_investimento().
    Ele saía da função ainda dentro da Ordem de Compra/Atribuições.
    Na iteração seguinte, resetar_me53n() redigitava "ME53N" no campo
    /nfe — mas esse campo só existe na janela principal do ME53N, e
    estando dentro da Ordem de Compra esse campo tem outro propósito
    (pode nem existir, ou pertencer a outra transação). O Enter seguinte
    então executava uma ação dentro da tela errada — daí o "abre uma
    coisa nada a ver com o processo".

    Esta função força a saída de qualquer profundidade através de F3
    repetidos ANTES de tentar redigitar a transação.
    """
    # Fecha qualquer modal residual primeiro (popup de erro, confirmação, etc.)
    fechar_modais(session)

    # Pressiona F3 várias vezes: cobre Atribuições -> Ordem -> PR -> ME53N inicial
    # Cada chamada é protegida individualmente — se o F3 não fizer efeito
    # (ex: já está na tela inicial), o SAP simplesmente ignora, sem erro.
    for _ in range(4):
        try:
            session.findById("wnd[0]").sendVKey(3)  # F3 = Voltar
            time.sleep(0.4)
        except:
            pass
        fechar_modais(session)  # cada F3 pode abrir um popup de confirmação

    # Reforço final: redigita ME53N no campo de transação.
    # Mesmo já estando na tela inicial, isso não causa efeito colateral
    # e garante 100% que estamos na transação certa antes do Shift+F5.
    try:
        session.findById("wnd[0]/tbar[0]/okcd").text = "ME53N"
        session.findById("wnd[0]").sendVKey(0)
    except:
        pass

    # Aguarda a tela inicial confirmando que o botão da toolbar existe
    wait_for_element(session, "wnd[0]/tbar[1]/btn[6]", timeout=10)
    time.sleep(0.5)


# =========================================================
# RESETAR ME53N PARA TELA INICIAL (SEM PR ABERTA)
# =========================================================
def resetar_me53n(session):
    """
    Wrapper mantido por compatibilidade — delega para voltar_para_inicio_me53n,
    que agora faz a saída de qualquer profundidade de tela (e não apenas
    redigita a transação, que sozinho não era suficiente).
    """
    voltar_para_inicio_me53n(session)


# =========================================================
# ABRIR PR NA ME53N
# =========================================================
def abrir_pr(session, pr):
    """
    Abre uma PR na ME53N — fluxo confirmado pelo script gravado no SAP:
      1. sendVKey(17) = Shift+F5 — abre popup "Selecionar documento"
      2. Digita o número da PR no campo ctxtMEPO_SELECT-BANFN
      3. sendVKey(0) = Enter para carregar a PR

    Esse sendVKey(17) só funciona na tela INICIAL do ME53N — garantida
    pelo resetar_me53n() chamado antes desta função no loop principal.
    """
    # Confirma que a tela inicial do ME53N está pronta antes do Shift+F5
    wait_for_element(session, "wnd[0]/tbar[1]/btn[6]", timeout=10)

    # Abre o popup "Selecionar documento" (Shift+F5)
    session.findById("wnd[0]").sendVKey(17)

    # Digita o número da PR no campo do popup
    campo = wait_for_element(
        session,
        "wnd[1]/usr/subSUB0:SAPLMEGUI:0003/ctxtMEPO_SELECT-BANFN",
        timeout=10
    )
    campo.text = pr

    # Enter para carregar a PR
    session.findById("wnd[1]").sendVKey(0)
    time.sleep(1.5)  # aguarda a PR carregar completamente


# =========================================================
# VERIFICAR INVESTIMENTO (ELEMENTO PEP)
# =========================================================
def verificar_investimento(session):
    """
    Navega até a Ordem de Compra vinculada à PR e verifica o Elemento PEP.

    Fluxo replicando o processo manual descrito:
      1. Desce 2-3x com seta para baixo na tabela de itens da PR
         (posiciona o cursor no item correto antes do F2)
      2. F2 para abrir a Ordem de Compra
      3. sendVKey(6) = Ctrl+→ para ir à aba "Atribuições"
      4. Verifica o Elemento PEP
      5. F3 para voltar — feito aqui dentro, e reforçado depois pelo
         voltar_para_inicio_me53n() chamado no loop principal

    Retorna "Investimento" se PEP preenchido, "" caso contrário.
    """
    # ------------------------------------------------------------------
    # PASSO 1: Garante que a aba "Avaliação" da PR está ativa
    # ------------------------------------------------------------------
    try:
        aba_avaliacao = wait_for_element(
            session,
            "wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100"
            "/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301"
            "/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6",
            timeout=10
        )
        aba_avaliacao.select()
        time.sleep(0.5)
    except:
        pass  # aba já pode estar ativa

    # ------------------------------------------------------------------
    # PASSO 2: Localiza o campo da Ordem de Compra e posiciona o cursor
    # (equivalente ao "desce 2-3x com seta" — aqui feito via caretPosition
    # direto no campo já identificado, que tem o mesmo efeito de
    # posicionar o cursor na linha do item antes do F2)
    # ------------------------------------------------------------------
    campo_ordem = wait_for_element(
        session,
        "wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100"
        "/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301"
        "/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6"
        "/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100"
        "/subSUB1:SAPLMEACCTVI:1100/subKONTBLOCK:SAPLKACB:1101/lblCOBL-AUFNR",
        timeout=10
    )
    campo_ordem.setFocus()
    campo_ordem.caretPosition = 5

    # ------------------------------------------------------------------
    # PASSO 3: F2 para abrir a Ordem de Compra
    # ------------------------------------------------------------------
    session.findById("wnd[0]").sendVKey(2)
    time.sleep(1)

    # ------------------------------------------------------------------
    # PASSO 4: Ctrl+→ (sendVKey 6) para ir à aba "Atribuições" da Ordem
    # ------------------------------------------------------------------
    session.findById("wnd[0]").sendVKey(6)
    time.sleep(0.5)

    # ------------------------------------------------------------------
    # PASSO 5: Tenta ler o Elemento PEP
    # Campo ausente ou vazio = não é investimento
    # ------------------------------------------------------------------
    try:
        pep = wait_for_element(
            session,
            "wnd[0]/usr/tabsTABSTRIP_600/tabpBUT2"
            "/ssubAREA_FOR_601:SAPMKAUF:0601/subAREA1:SAPMKAUF:0315"
            "/ctxtCOAS-PSPEL",
            timeout=5
        ).text
        valor = "Investimento" if pep.strip() else ""
    except:
        valor = ""

    # ------------------------------------------------------------------
    # PASSO 6 (CORREÇÃO PRINCIPAL — estava faltando):
    # F3 para voltar da Ordem de Compra para a tela da PR.
    # Sem isso, a função terminava ainda dentro da Ordem, e a próxima
    # chamada de resetar_me53n() não conseguia se recuperar corretamente
    # porque o campo /nfe não tem o mesmo comportamento em todas as telas.
    # ------------------------------------------------------------------
    try:
        session.findById("wnd[0]").sendVKey(3)  # F3 — sai da Ordem, volta pra PR
        time.sleep(0.5)
    except:
        pass

    return valor


# =========================================================
# UI — BARRA DE PROGRESSO COM CANCELAMENTO
# =========================================================
class ProgressApp:
    def __init__(self, total):
        self.root = tk.Tk()
        self.root.title("Analisando PRs tipo J")
        self.root.geometry("420x210")

        self.total = total
        self.current = 0
        self.start_time = time.time()
        self.cancelled = False

        self.label = tk.Label(self.root, text="Iniciando...", font=("Arial", 12))
        self.label.pack(pady=10)

        self.progress = ttk.Progressbar(self.root, length=320, maximum=total)
        self.progress.pack(pady=10)

        self.time_label = tk.Label(self.root, text="")
        self.time_label.pack()

        self.cancel_btn = tk.Button(
            self.root, text="Cancelar", command=self.cancel, bg="red", fg="white"
        )
        self.cancel_btn.pack(pady=10)

        # Botão "Concluído" começa oculto — aparece ao terminar
        self.done_btn = tk.Button(
            self.root, text="Concluído ✅", command=self.close, bg="green", fg="white"
        )
        self.done_btn.pack(pady=5)
        self.done_btn.pack_forget()

    def update(self, pr=""):
        if self.cancelled:
            return
        self.current += 1
        self.progress["value"] = self.current

        elapsed = time.time() - self.start_time
        avg = elapsed / self.current if self.current else 0
        restante = avg * (self.total - self.current)

        texto = f"{self.current}/{self.total}"
        if pr:
            texto += f"  —  PR {pr}"
        self.label.config(text=texto)
        self.time_label.config(text=f"Tempo restante: {int(restante)}s")
        self.root.update_idletasks()

    def cancel(self):
        self.cancelled = True
        self.label.config(text="Cancelando...")

    def finish(self, error=False):
        if error:
            self.label.config(text="Erro na execução ❌")
        elif self.cancelled:
            self.label.config(text="Cancelado ⚠️")
        else:
            self.label.config(text="Finalizado ✅")
        self.time_label.config(text="")
        self.cancel_btn.pack_forget()
        self.done_btn.pack()

    def close(self):
        self.root.destroy()


# =========================================================
# PROCESSO PRINCIPAL
# =========================================================
def run_process(app):

    pythoncom.CoInitialize()

    SALVAR_A_CADA = 10  # auto-save a cada N PRs processadas
    prs_desde_save = 0

    try:
        # ----------------------------------------------------------
        # INICIALIZAÇÃO: leitura do Excel e conexão SAP
        # ----------------------------------------------------------
        try:
            arquivo_entrada = r"C:\Users\MDEMOUR\Downloads\260504_Análise PRs Fórum 2.xlsx"
            arquivo_saida   = r"C:\Users\MDEMOUR\Downloads\260504_saida.xlsx"

            print("Lendo Excel...")
            df = pd.read_excel(arquivo_entrada)

            wb = load_workbook(arquivo_entrada)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]

            # CORREÇÃO v1: só insere coluna "Comentários" se ainda não existir.
            # Sem essa verificação, cada execução criava uma coluna nova,
            # deslocando todas as colunas seguintes e corrompendo o Excel.
            if "Comentários" not in headers:
                col_cc = headers.index("Centro de Custo") + 1
                ws.insert_cols(col_cc + 1)
                ws.cell(row=1, column=col_cc + 1).value = "Comentários"
                headers = [cell.value for cell in ws[1]]  # recarrega após inserção

            col_comentarios = headers.index("Comentários") + 1  # base-1 para openpyxl

            print("Conectando SAP...")
            session = get_sap_session()

            # Abre a transação ME53N e aguarda carregar
            session.findById("wnd[0]/tbar[0]/okcd").text = "ME53N"
            session.findById("wnd[0]").sendVKey(0)
            time.sleep(1)

        except Exception as e:
            print("\nERRO NA INICIALIZAÇÃO:\n", e)
            app.finish(error=True)
            return

        # ----------------------------------------------------------
        # LOOP PRINCIPAL: processa cada linha do Excel
        # ----------------------------------------------------------
        print("Processando PRs...")

        for index, row in df.iterrows():

            if app.cancelled:
                print("Processamento cancelado pelo usuário.")
                break

            tipo_io = str(row["Tipo_IO"]).strip()

            # CORREÇÃO v1: filtro mais robusto que o regex original ^\d{3}J$
            # Aceita qualquer Tipo_IO que termine em "J" precedido de dígitos
            # (ex: "540J", "564J", "0001J" — todos passam).
            # NaN vira a string "nan" pelo str(), que não termina em J → pulado.
            eh_tipo_j = tipo_io.upper().endswith("J") and tipo_io[:-1].isdigit()

            # CORREÇÃO v1: linha_excel derivada do índice pandas, não de contador manual.
            # Contador manual desincronizava quando linhas eram puladas.
            linha_excel = index + 2  # +2: cabeçalho na linha 1, índice pandas começa em 0

            if not eh_tipo_j:
                app.update()
                continue

            pr = str(row["Nº PR"]).strip()
            # Remove ".0" caso a coluna tenha sido lida como float pelo pandas
            if pr.endswith(".0"):
                pr = pr[:-2]

            try:
                # ------------------------------------------------------
                # CORREÇÃO v2 (causa raiz 'virtual key not enabled'):
                # Garante que estamos na tela INICIAL do ME53N antes de
                # chamar sendVKey(17). Sem isso, a primeira PR funciona,
                # mas todas as seguintes falham porque o SAP ainda está
                # na tela da PR anterior, onde Shift+F5 não é válido.
                # ------------------------------------------------------
                resetar_me53n(session)

                # Abre a PR na ME53N
                abrir_pr(session, pr)

                # Navega até a Ordem de Compra e verifica Elemento PEP
                valor = verificar_investimento(session)

                print(f"  PR {pr} → {valor if valor else 'Despesa'}")

            except Exception as e:
                print(f"  ERRO PR {pr} (linha {linha_excel}): {e}")
                valor = "Erro"

                # Tenta recuperar o estado do SAP para continuar o loop
                # sem travar nas PRs seguintes
                try:
                    resetar_me53n(session)
                except:
                    pass

            # Grava resultado na coluna Comentários
            ws.cell(row=linha_excel, column=col_comentarios).value = valor

            # Auto-save periódico: evita perda de dados em cancelamentos
            prs_desde_save += 1
            if prs_desde_save >= SALVAR_A_CADA:
                wb.save(arquivo_saida)
                prs_desde_save = 0
                print(f"  [Auto-save] checkpoint salvo após PR {pr}")

            app.update(pr=pr)

        # ----------------------------------------------------------
        # FINALIZAÇÃO
        # ----------------------------------------------------------
        print("Salvando arquivo final...")
        wb.save(arquivo_saida)
        print(f"Concluído. Arquivo salvo em:\n  {arquivo_saida}")

        app.finish()

    except Exception as e:
        print("\nERRO GERAL:\n", e)
        app.finish(error=True)

    finally:
        pythoncom.CoUninitialize()


# =========================================================
# ENTRADA DO PROGRAMA
# =========================================================
arquivo_entrada = r"C:\Users\MDEMOUR\Downloads\260504_Análise PRs Fórum 2.xlsx"

df_temp = pd.read_excel(arquivo_entrada)
total = len(df_temp)

app = ProgressApp(total)

# Processo SAP roda em thread separada para não bloquear a UI tkinter
threading.Thread(target=run_process, args=(app,), daemon=True).start()

app.root.mainloop()
